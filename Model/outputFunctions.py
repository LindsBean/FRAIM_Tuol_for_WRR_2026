import pandas as pd
import numpy as np
from openpyxl import load_workbook
# import os
# os.chdir(os.path.dirname(os.path.abspath(__file__))) 

# funtion
def saveToSheet(df_list, sheets, file_name, spaces, decision_str=''):
    book = load_workbook(file_name)
    writer = pd.ExcelWriter(file_name,engine='openpyxl') 
    row = 2
    for dataframe in df_list:
        writer.book = book
        dataframe.to_excel(writer,sheet_name=sheets,startrow=(row+1) , startcol=0)   
        # worksheet = writer.sheets[sheets] 
        # worksheet.write(row, 0, dataframe.name)
        row = row + len(dataframe.index) + spaces + 1
    worksheet= book.get_sheet_by_name(sheets) 
    cell = worksheet.cell(1,1)
    cell.value = decision_str
    writer.save()
    writer.close()
    print('Save Successful [1/2]')

def addSheetFromWorkbook(temp_path, target_path, sheetname):
    import xlwings as xw
    path1 = temp_path
    path2 = target_path

    wb1 = xw.Book(path1)
    wb2 = xw.Book(path2)

    ws1 = wb1.sheets[sheetname]
    ws1.api.copy_worksheet(after_=wb2.sheets[-1].api)
    wb2.save()
    wb2.app.quit()
    print('Save Successful [2/2]')




# array=[[0,3],[4,5]]
# df = pd.DataFrame(array)
# df1 = pd.DataFrame(array)
# df2 = pd.DataFrame(array)
# # list of dataframes
# dfs = [df,df1,df2]

# # run function
# multiple_dfs(dfs, 'Validation', 'test1.xlsx', 1)

# book = load_workbook(path)
# writer = pd.ExcelWriter(path, engine = 'openpyxl')
# writer.book = book

# df.to_excel(writer, sheet_name=updateMo, startrow=2)

# worksheet = writer.sheets[updateMo]
# worksheet['A1'] = str(WY) + ' ' + updateMo + ' Update'
# worksheet['B2'] = 'Optimal Volume by Season (+ carryover)'
# worksheet['G2'] = 'Performance Function Results '

# writer.save()
# writer.close()
# print('Save Successful')

